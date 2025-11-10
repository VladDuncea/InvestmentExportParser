import sys
import csv
from openpyxl import Workbook, load_workbook, utils
import datetime
import uuid
from forex_python.converter import CurrencyRates
import json
import os.path

c = CurrencyRates()
useFXRates = False

# TODO cleanup a lot pls
fxRateCache = {}
if os.path.isfile("ratesCache.txt"):
    with open("ratesCache.txt", "r") as file:
        fxRateCache = json.loads(file.read())


ignore = True
etoroList = [
    1121381748,
    1128698036,
    1138197295,
    1138698607,
    1174312587,
    1204353973,
    1204363261,
    1177887965,
    1357911361,
    1365314951,
    2404349632,
    2443378060,
    1213782880,
    2592641833,
]


def etoroDateToDateTime(value: str) -> datetime.datetime:
    return datetime.datetime.strptime(value, "%d/%m/%Y %H:%M:%S")


def extractDateFromDateTime(date: datetime.datetime) -> datetime.datetime:
    return datetime.datetime(year=date.year, month=date.month, day=date.day)


def compareDates(date1: datetime.datetime, date2: datetime.datetime):
    return date1.strftime("%Y-%m-%d") == date2.strftime("%Y-%m-%d")


class InvestmentParser:
    def __init__(self, filePath: str, type: str):
        self.filePath = filePath
        self.type = type
        # Init result xls file
        self.initResultXls()
        self.cacheDict = {
            "dividends": [],
            "deposits": [],
            "sales": [],
            "taxes_comissions": [],
            "intermediarySales": {},
            "deltaRows": [],
        }

    def getFxRate(self, transactDate: datetime.datetime):
        dateString = transactDate.strftime("%Y_%m_%d")
        if dateString not in fxRateCache:
            if not useFXRates:
                return 1
            print(f"Loading FX Rate for {dateString}")
            fxRateCache[dateString] = c.get_rate("USD", "RON", transactDate)
            print(f"Loaded FX Rate for {dateString}")
        return fxRateCache[dateString]

    def initResultXls(self):
        self.resultXls = Workbook()
        sheet = self.resultXls.active
        # rename initial sheet
        sheet.title = "Dividends"
        # create sheet for deposits
        self.resultXls.create_sheet("Deposits")
        # create sheet for sales
        self.resultXls.create_sheet("Sales")
        # create sheet for taxes/comissions
        self.resultXls.create_sheet("Taxes+Comissions")

    def parse(self):
        if self.type == "xtb":
            self.parseXtb()
        elif self.type == "etoro":
            self.parseEtoro()
        elif self.type == "revolut":
            self.parseRevolut()
        else:
            print(
                f"ERR: Wrong type of file, expected 'xtb', 'etoro', 'revolut', got '{self.type}'\n"
            )

    ################# REVOLUT ##########################
    def parseRevolut(self):
        # Define variable to load the dataframe
        excelFile = load_workbook(self.filePath)

        # Define variable to read sheet
        sheetMain = excelFile.active

        # Iterate over the rows and col
        for row in sheetMain.iter_rows(2, sheetMain.max_row):
            self.handleRevolutMainSheetRow(row)

        self.exportResult("revolut")

    def handleRevolutMainSheetRow(self, row):
        try:
            transactDate = extractDateFromDateTime(
                datetime.datetime.fromisoformat(row[0].value.replace("Z", ""))
            )
            transactFullDate = datetime.datetime.fromisoformat(
                row[0].value.replace("Z", "")
            )
            transactSymbol = row[1].value  # only some transact types have it
            transactType = row[2].value
            transactQuantity = row[3].value  # only some transact types have it
            pricePerShare = row[4].value  # only some transact types have it
            totalValue = float(row[5].value)  # all have total value
            fxRate = float(row[7].value)  # fx rate -> to ron
        except:
            # maybe out of data range
            return

        # # parse total value as it is an accounting value
        # if totalValueStr[0] == "(":
        #     # negative value
        #     totalValue = -totalValueStr.replace("(", "").replace(")", "")
        # else:
        #     totalValue = totalValueStr

        if transactType in ["DIVIDEND", "DIVIDEND TAX (CORRECTION)"]:
            self.cacheDict["dividends"].append(
                {
                    "date": transactDate,
                    "fullDate": transactFullDate,
                    "company": transactSymbol,
                    "value": totalValue,
                }
            )
            self.cacheDict["deltaRows"].append(
                {
                    "action": "DIVIDEND",
                    "amount": "",
                    "fullDate": transactFullDate,
                    "company": transactSymbol,
                    "value": totalValue,
                    "type": "STOCK",
                }
            )
        elif transactType in ["CASH TOP-UP", "CASH WITHDRAWAL"]:
            if (
                len(self.cacheDict["deposits"]) > 0
                and self.cacheDict["deposits"][-1]["date"] == transactDate
            ):
                self.cacheDict["deposits"][-1]["value"] += totalValue
                self.cacheDict["deposits"][-1]["value_ron"] += (1 / fxRate) * totalValue
            else:
                self.cacheDict["deposits"].append(
                    {
                        "date": transactDate,
                        "value_ron": (1 / fxRate) * totalValue,
                        "value": totalValue,
                    }
                )
            self.cacheDict["deltaRows"].append(
                {
                    "action": (
                        "DEPOSIT" if transactType == "CASH TOP-UP" else "WITHDRAW"
                    ),
                    "amount": (
                        totalValue if transactType == "CASH TOP-UP" else -totalValue
                    ),
                    "fullDate": transactFullDate,
                    "company": "USD",
                    "value": "",
                    "type": "FIAT",
                }
            )
        elif transactType in ["CUSTODY FEE"]:
            self.cacheDict["taxes_comissions"].append(
                {
                    "date": transactDate,
                    "value": totalValue,
                    "type": "monthly fee",
                    "moreInfo": transactType,
                }
            )
            self.cacheDict["deltaRows"].append(
                {
                    "action": "WITHDRAW",
                    "amount": -totalValue,
                    "fullDate": transactFullDate,
                    "company": "USD",
                    "value": "",
                    "type": "FIAT",
                }
            )
        elif transactType in [
            "BUY - MARKET",
            "STOCK SPLIT",
        ]:  # stock splits tell the delta after the split (unfortunatley negative deltas seem to be wrongly reported by revolut as positive)
            if transactSymbol not in self.cacheDict["intermediarySales"]:
                self.cacheDict["intermediarySales"][transactSymbol] = {
                    "count": 0,
                    "value": 0,
                    "firstDate": transactDate,
                }
            self.cacheDict["intermediarySales"][transactSymbol][
                "count"
            ] += transactQuantity
            self.cacheDict["intermediarySales"][transactSymbol]["value"] += totalValue
            if transactType in ["BUY - MARKET"]:
                self.cacheDict["deltaRows"].append(
                    {
                        "action": "BUY",
                        "amount": transactQuantity,
                        "fullDate": transactFullDate,
                        "company": transactSymbol,
                        "value": totalValue,
                        "type": "STOCK",
                    }
                )
        elif transactType in ["SELL - MARKET"]:
            # we have to do some calculations
            knownInfo = self.cacheDict["intermediarySales"][transactSymbol]
            # calculate open value as an average of all the open positions
            openValue = knownInfo["value"] / knownInfo["count"] * transactQuantity
            # update known values
            self.cacheDict["intermediarySales"][transactSymbol][
                "count"
            ] -= transactQuantity
            self.cacheDict["intermediarySales"][transactSymbol]["value"] -= openValue
            # add info
            self.cacheDict["sales"].append(
                {
                    "dateOpen": self.cacheDict["intermediarySales"][transactSymbol][
                        "firstDate"
                    ],  # this cannot be decided from revolut, so mark with empty
                    "dateClose": transactDate,
                    "company": transactSymbol,
                    "openValue": openValue,
                    "closeValue": totalValue,
                }
            )
            self.cacheDict["deltaRows"].append(
                {
                    "action": "SELL",
                    "amount": transactQuantity,
                    "fullDate": transactFullDate,
                    "company": transactSymbol,
                    "value": totalValue,
                    "type": "STOCK",
                }
            )
        else:
            print(f"WARN: Unknown transaction type: {transactType}")

    ################# END REVOLUT ######################

    ################# XTB ##########################

    def parseXtb(self):
        # Define variable to load the dataframe
        excelFile = load_workbook(self.filePath)

        # Define variable to read sheet
        sheetCashOp = excelFile["CASH OPERATION HISTORY"]

        # Iterate over the rows and col
        for row in sheetCashOp.iter_rows(12, sheetCashOp.max_row):
            self.handleXtbCashHistRow(row)

        # Define variable to read sheet
        sheetClosedOp = excelFile["CLOSED POSITION HISTORY"]

        # Iterate over the rows and col
        for row in sheetClosedOp.iter_rows(14, sheetClosedOp.max_row):
            self.handleXtbClosedOpRow(row)

        ## Dividends for delta
        for row in self.cacheDict["dividends"]:
            self.cacheDict["deltaRows"].append(
                {
                    "action": "DIVIDEND" if row["company"] != "DOBANDA" else "DEPOSIT",
                    "amount": "" if row["company"] != "DOBANDA" else row["value"],
                    "fullDate": row["fullDate"],
                    "company": (
                        deltaTickerHelper(row["company"])
                        if row["company"] != "DOBANDA"
                        else "USD"
                    ),
                    "value": row["value"] if row["company"] != "DOBANDA" else "",
                    "type": (
                        deltaCompanyHelper(row["company"])
                        if row["company"] != "DOBANDA"
                        else "FIAT"
                    ),
                }
            )

        self.exportResult("xtb")

    def handleXtbClosedOpRow(self, row):
        try:
            transactDateOpen = extractDateFromDateTime(row[5].value)
            transactDateClose = extractDateFromDateTime(row[7].value)
            transactFullDate = row[7].value
            transactSymbol = row[2].value
            volume = row[4].value
            openValue = row[11].value
            closeValue = row[12].value
        except:
            # maybe out of data range
            return

        self.cacheDict["sales"].append(
            {
                "dateOpen": transactDateOpen,
                "dateClose": transactDateClose,
                "company": transactSymbol,
                "openValue": openValue,
                "closeValue": closeValue,
            }
        )
        companyTicker = deltaTickerHelper(transactSymbol)

        # if (
        #     self.cacheDict["deltaRows"][-1]
        #     and self.cacheDict["deltaRows"][-1]["fullDate"] == transactFullDate
        #     and self.cacheDict["deltaRows"][-1]["company"] == companyTicker
        # ):
        #     transactFullDate += datetime.timedelta(0, 1)

        self.cacheDict["deltaRows"].append(
            {
                "action": "SELL",
                "amount": volume,
                "fullDate": transactFullDate,
                "company": deltaTickerHelper(transactSymbol),
                "value": closeValue,
                "type": deltaCompanyHelper(transactSymbol),
                "comment": uuid.uuid4(),
            }
        )

    def handleXtbCashHistRow(self, row):
        try:
            transactType = row[2].value
            transactDate = extractDateFromDateTime(row[3].value)
            transactFullDate = row[3].value
            transactComment = row[4].value
            transactSymbol = row[5].value  # not all rows have a symbol
            if transactType in [
                "Free-funds Interest",
                "Free-funds Interest Tax",
            ]:
                transactSymbol = "DOBANDA"
            if transactType == "spin-off":
                transactSymbol += " (spin-off)"
            value = row[6].value
        except:
            # maybe out of data range
            return

        if transactType in [
            "Dividend",
            "DIVIDENT",
            "spin-off",
            "Withholding tax",
            "Withholding Tax",
            "Stamp duty",
            "Stamp Duty",
            "Free-funds Interest",
            "Free-funds Interest Tax",
            "Impozitul reținut",
        ]:
            if (
                len(self.cacheDict["dividends"]) > 0
                and self.cacheDict["dividends"][-1]["date"] == transactDate
                and self.cacheDict["dividends"][-1]["company"] == transactSymbol
            ):
                self.cacheDict["dividends"][-1]["value"] += value
            else:
                self.cacheDict["dividends"].append(
                    {
                        "date": transactDate,
                        "fullDate": transactFullDate,
                        "company": transactSymbol,
                        "value": value,
                    }
                )
        elif transactType in ["Deposit", "Depunere", "deposit"]:
            fxRate = self.getFxRate(transactFullDate)
            if (
                len(self.cacheDict["deposits"]) > 0
                and self.cacheDict["deposits"][-1]["date"] == transactDate
            ):
                self.cacheDict["deposits"][-1]["value"] += value
                self.cacheDict["deposits"][-1]["value_ron"] += fxRate * value
            else:
                self.cacheDict["deposits"].append(
                    {"date": transactDate, "value_ron": fxRate * value, "value": value}
                )
            self.cacheDict["deltaRows"].append(
                {
                    "action": (
                        "DEPOSIT"
                        if transactType in ["Deposit", "Depunere"]
                        else "WITHDRAW"
                    ),
                    "amount": (
                        value if transactType in ["Deposit", "Depunere"] else -value
                    ),
                    "fullDate": transactFullDate,
                    "company": "USD",
                    "value": "",
                    "type": "FIAT",
                }
            )
        elif transactType in ["tax RO", "SEC fee", "Sec Fee"]:
            self.cacheDict["taxes_comissions"].append(
                {
                    "date": transactDate,
                    "value": value,
                    "type": transactType,
                    "moreInfo": transactComment,
                }
            )
        elif transactType in [
            "Stocks/ETF purchase",
            "Acțiuni/Cumpărare ETF",
            "Stock purchase",
        ]:
            self.cacheDict["deltaRows"].append(
                {
                    "action": "BUY",
                    "amount": transactComment.split(" ")[2].split("/")[0],
                    "fullDate": transactFullDate,
                    "company": deltaTickerHelper(transactSymbol),
                    "value": -value,
                    "type": deltaCompanyHelper(transactSymbol),
                }
            )
        elif transactType in [
            "Profit/Loss",
            "Stocks/ETF sale",
            "Vânzare acțiuni /ETF-uri",
            "Profit/Pierdere",
            "Stock sale",
            "close trade",
        ]:
            # nothing to do yet
            pass
        else:
            print(f"WARN: Unknown transaction type: {transactType}")

    ################# END XTB #######################

    ################# ETORO ##########################

    def parseEtoro(self):
        # Define variable to load the dataframe
        excelFile = load_workbook(self.filePath)

        # Define variable to read sheet
        sheetAccActivity = excelFile["Account Activity"]

        # Iterate over the rows and col
        for row in sheetAccActivity.iter_rows(2, sheetAccActivity.max_row):
            self.handleEtoroAccActivityRow(row)

        # Define variable to read sheet
        sheetClosedOp = excelFile["Closed Positions"]

        # Iterate over the rows and col
        for row in sheetClosedOp.iter_rows(2, sheetClosedOp.max_row):
            self.handleEtoroClosedOpRow(row)

        self.exportResult("etoro")

    def handleEtoroClosedOpRow(self, row):
        try:
            transactID = row[0].value
            if (
                transactID
                and transactID != "-"
                and (int(transactID) in etoroList) == ignore
            ):
                return
            transactDateOpen = extractDateFromDateTime(
                etoroDateToDateTime(row[5].value)
            )
            transactDateClose = extractDateFromDateTime(
                etoroDateToDateTime(row[6].value)
            )
            openValue = float(row[3].value)
            closeValue = openValue + float(row[10].value)
            transactSymbol = self.cacheDict["intermediarySales"][transactID]
            rolloverDivFees = "" if row[17].value == 0 else f" ({row[17].value})"
            copyFrom = "" if row[18].value == "-" else f" ({row[18].value})"
        except:
            # maybe out of data range
            print("WARN: An exception occurred. ID:" + transactID)
            return

        self.cacheDict["sales"].append(
            {
                "dateOpen": transactDateOpen,
                "dateClose": transactDateClose,
                "company": transactSymbol + rolloverDivFees + copyFrom,
                "openValue": openValue,
                "closeValue": closeValue,
            }
        )

    def handleEtoroAccActivityRow(self, row):
        try:
            transactID = row[8].value  # not all rows have transact ID
            if (
                transactID
                and transactID != "-"
                and (int(transactID) in etoroList) == ignore
            ):
                return
            transactType = row[1].value
            transactDate = extractDateFromDateTime(etoroDateToDateTime(row[0].value))
            # transactComment = row[4].value
            transactSymbol = (
                row[2].value.split("/")[0] if row[2].value else ""
            )  # not all rows have a symbol
            value = row[3].value
        except:
            print("WARN: An exception occurred")
            # maybe out of data range
            return
        if transactType in ["Dividend"]:
            if (
                len(self.cacheDict["dividends"]) > 0
                and self.cacheDict["dividends"][-1]["date"] == transactDate
                and self.cacheDict["dividends"][-1]["company"] == transactSymbol
            ):
                self.cacheDict["dividends"][-1]["value"] += value
            else:
                self.cacheDict["dividends"].append(
                    {
                        "date": transactDate,
                        "company": self.cacheDict["intermediarySales"][transactID],
                        "value": value,
                    }
                )
        elif transactType in ["Deposit", "Interest Payment"]:
            fxRate = self.getFxRate(transactDate)
            if (
                len(self.cacheDict["deposits"]) > 0
                and self.cacheDict["deposits"][-1]["date"] == transactDate
            ):
                self.cacheDict["deposits"][-1]["value"] += value
                self.cacheDict["deposits"][-1]["value_ron"] += fxRate * value
            else:
                self.cacheDict["deposits"].append(
                    {
                        "date": transactDate,
                        "value_ron": fxRate * value,
                        "value": value,
                    }  # value_ron for consistency, cant be calculated
                )
        elif transactType in ["Open Position", "Position closed"]:
            self.cacheDict["intermediarySales"][transactID] = transactSymbol
        elif transactType in ["Overnight fee"]:
            self.cacheDict["taxes_comissions"].append(
                {
                    "date": transactDate,
                    "value": value,
                    "type": transactType,
                    "moreInfo": "",
                }
            )
        elif transactType in [
            "Start Copy",
            "Stop Copy",
            "corp action: Split",
            "Adjustment",
        ]:
            pass
        else:
            print(f"WARN: Unknown transaction type: {transactType}")

    ################# END ETORO #######################

    def exportResult(self, filePrefix: str):
        # Dividend sheet
        sheet = self.resultXls["Dividends"]
        sheet.append(["Date", "Company", "Value"])
        for idx, row in enumerate(self.cacheDict["dividends"], 2):
            sheet.append([row["date"], row["company"], row["value"]])
            sheet[f"A{idx}"].number_format = "dd-mm-yy"
            sheet[f"C{idx}"].number_format = (
                '_([$$-en-US]* #,##0.00_);_([$$-en-US]* (#,##0.00);_([$$-en-US]* "-"??_);_(@_)'
            )

        # Deposits sheet
        sheet = self.resultXls["Deposits"]
        sheet.append(["Date", "Value"])
        for idx, row in enumerate(self.cacheDict["deposits"], 2):
            sheet.append([row["date"], row["value_ron"], row["value"]])
            sheet[f"A{idx}"].number_format = "dd-mm-yy"
            sheet[f"B{idx}"].number_format = (
                '_-* #,##0.00 [$lei-ro-RO]_-;-* #,##0.00 [$lei-ro-RO]_-;_-* "-"?? [$lei-ro-RO]_-;_-@_-'
            )
            sheet[f"C{idx}"].number_format = (
                '_([$$-en-US]* #,##0.00_);_([$$-en-US]* (#,##0.00);_([$$-en-US]* "-"??_);_(@_)'
            )

        # Sales sheet
        sheet = self.resultXls["Sales"]
        sheet.append(
            [
                "Company",
                "Open Date",
                "Sell Date",
                "Buy Value",
                "Sell Value",
                "Profit",
            ]
        )
        for idx, row in enumerate(self.cacheDict["sales"], 2):
            sheet.append(
                [
                    row["company"],
                    row["dateOpen"],
                    row["dateClose"],
                    row["openValue"],
                    row["closeValue"],
                    row["closeValue"] - row["openValue"],
                ]
            )
            sheet[f"B{idx}"].number_format = "dd-mm-yy"
            sheet[f"C{idx}"].number_format = "dd-mm-yy"
            sheet[f"D{idx}"].number_format = (
                '_([$$-en-US]* #,##0.00_);_([$$-en-US]* (#,##0.00);_([$$-en-US]* "-"??_);_(@_)'
            )
            sheet[f"E{idx}"].number_format = (
                '_([$$-en-US]* #,##0.00_);_([$$-en-US]* (#,##0.00);_([$$-en-US]* "-"??_);_(@_)'
            )
            sheet[f"F{idx}"].number_format = (
                '_([$$-en-US]* #,##0.00_);_([$$-en-US]* (#,##0.00);_([$$-en-US]* "-"??_);_(@_)'
            )

        # Taxes and comissions sheet
        sheet = self.resultXls["Taxes+Comissions"]
        sheet.append(["Reason", "Date", "Value", "Comment"])
        for idx, row in enumerate(self.cacheDict["taxes_comissions"], 2):
            sheet.append([row["type"], row["date"], row["value"], row["moreInfo"]])
            sheet[f"B{idx}"].number_format = "dd-mm-yy"
            sheet[f"C{idx}"].number_format = (
                '_([$$-en-US]* #,##0.00_);_([$$-en-US]* (#,##0.00);_([$$-en-US]* "-"??_);_(@_)'
            )

        # export parse result
        self.resultXls.save(
            f"exportFiles/{filePrefix}_investments_{datetime.datetime.now().strftime('%Y_%m_%d')}.xlsx"
        )

        ## DELTA
        with open(f"exportFiles/{filePrefix}_delta.csv", "w", newline="") as csvfile:
            csvWriter = csv.writer(csvfile)
            csvWriter.writerow(
                [
                    "Date",
                    "Way",
                    "Base amount",
                    "Base currency (name)",
                    "Base type",
                    "Quote amount",
                    "Quote currency",
                    "Exchange",
                    "Sent/Received from",
                    "Sent to",
                    "Fee amount",
                    "Fee currency (name)",
                    "Broker",
                    "Notes",
                ]
            )
            for row in self.cacheDict["deltaRows"]:
                csvWriter.writerow(
                    [
                        row["fullDate"].strftime("%Y-%m-%d %H:%M:%S.%f+00:00"),
                        row["action"],
                        row["amount"],
                        row["company"],
                        row["type"],
                        row["value"],
                        "USD",
                        "",
                        "",
                        "",
                        "",
                        "",
                        filePrefix,
                        row["comment"] if "comment" in row else "",
                    ]
                )


################# DELTA #######################


def deltaTickerHelper(ticker):
    parts = ticker.split(".")
    if parts[0] == "GOOGC":
        return "GOOG"
    if parts[1] == "UK":
        return parts[0] + ".L"
    if parts[1] == "DE":
        return parts[0] + ".DE"
    return parts[0]


def deltaCompanyHelper(ticker):
    parts = ticker.split(".")
    if parts[0] in ["CSPX", "CNDX", "IWDA", "RBOT", "IQQH", "EUNL"]:
        return "FUND"
    else:
        return "STOCK"


################# END DELTA ###################


# takes arguments form command line, expects 2 args, the path of the excel file and the type of import (xtb or etoro or revolut)
def main(args):
    if len(args) != 3:
        print(
            f"ERR: Wrong number of parameters! Got {len(args)-1}, but expecting 2 parameters: <path-to-excel> <type-of-import>\n"
        )

    # Params
    filePath = args[1]
    type = args[2]
    # create the class
    investmentParser = InvestmentParser(filePath, type)
    investmentParser.parse()


main(sys.argv)

# TODO cleanup a lot pls
with open("ratesCache.txt", "w") as file:
    file.write(json.dumps(fxRateCache))
